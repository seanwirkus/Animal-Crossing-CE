#!/usr/bin/env python3
"""
Convert slices of docs/acnh.xlsx into runtime-friendly data.

Usage:
    source .venv/bin/activate  # ensure openpyxl is available
    python scripts/import_acnh.py
"""
from __future__ import annotations

import argparse
import json
import math
import re
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

WORKBOOK_PATH = Path("docs/acnh.xlsx")
OUTPUT_JSON = Path("data/generated/items.json")
OUTPUT_LUA = Path("src/shared/Data/GeneratedItems.luau")

CATEGORY_SHEETS: Dict[str, str] = {
    "Fish": "Fish",
    "Insects": "Bug",
}

MONTH_COLUMNS: Sequence[Tuple[str, int]] = (
    ("NH Jan", 1),
    ("NH Feb", 2),
    ("NH Mar", 3),
    ("NH Apr", 4),
    ("NH May", 5),
    ("NH Jun", 6),
    ("NH Jul", 7),
    ("NH Aug", 8),
    ("NH Sep", 9),
    ("NH Oct", 10),
    ("NH Nov", 11),
    ("NH Dec", 12),
)

MONTH_TO_SEASON = {
    1: "Winter",
    2: "Winter",
    3: "Spring",
    4: "Spring",
    5: "Spring",
    6: "Summer",
    7: "Summer",
    8: "Summer",
    9: "Autumn",
    10: "Autumn",
    11: "Autumn",
    12: "Winter",
}

RARITY_THRESHOLDS: Sequence[Tuple[int, str]] = (
    (0, "Common"),
    (1_000, "Uncommon"),
    (3_000, "Rare"),
    (8_000, "UltraRare"),
)

LOCATION_KEYWORDS: Sequence[Tuple[str, str]] = (
    ("Sea (raining)", "LocationSeaRain"),
    ("Sea (rainy days)", "LocationSeaRain"),
    ("Sea (rainy day)", "LocationSeaRain"),
    ("Sea (rain)", "LocationSeaRain"),
    ("Sea (rainstorm)", "LocationSeaRain"),
    ("Sea", "LocationSea"),
    ("River (Mouth)", "LocationRiverMouth"),
    ("River (Clifftop)", "LocationRiverClifftop"),
    ("River (Clifftop) (rainy days)", "LocationRiverClifftopRain"),
    ("River (Clifftop) (rainy day)", "LocationRiverClifftopRain"),
    ("River (Clifftop)", "LocationRiverClifftop"),
    ("River (raining)", "LocationRiverRain"),
    ("River (rainy days)", "LocationRiverRain"),
    ("River (rainy day)", "LocationRiverRain"),
    ("River (rainstorm)", "LocationRiverRain"),
    ("River", "LocationRiver"),
    ("Pond", "LocationPond"),
    ("Lake", "LocationLake"),
    ("Pier", "LocationPier"),
    ("Harbor", "LocationHarbor"),
    ("Clifftop", "LocationCliff"),
    ("On trees (any kind)", "LocationTree"),
    ("On trees (hardwood)", "LocationTreeHardwood"),
    ("On trees (cedar)", "LocationTreeCedar"),
    ("On coconut trees", "LocationTreePalm"),
    ("On palm trees", "LocationTreePalm"),
    ("On tree stumps", "LocationTreeStump"),
    ("On flowers", "LocationFlower"),
    ("On white flowers", "LocationFlowerWhite"),
    ("On pink flowers", "LocationFlowerPink"),
    ("Flying near flowers", "LocationFlyingFlower"),
    ("Flying", "LocationAir"),
    ("Underground", "LocationUnderground"),
    ("On rocks", "LocationRock"),
    ("On beach rocks", "LocationBeachRock"),
    ("On rotten turnips", "LocationRottenTurnip"),
    ("On rotten fruit", "LocationRottenFruit"),
    ("On fallen fruit", "LocationGroundFruit"),
    ("Shaking trees", "LocationTreeShake"),
    ("Hitting rocks", "LocationRockHit"),
    ("Near trash", "LocationTrash"),
    ("On villagers", "LocationVillager"),
    ("On flowers (rain)", "LocationFlowerRain"),
    ("Snowballs", "LocationSnowball"),
    ("Hot spring", "LocationHotSpring"),
    ("Any river", "LocationRiver"),
    ("Any except rain", "LocationDry"),
    ("Any weather", "LocationAnyWeather"),
)

TIME_RANGE_SPLIT = re.compile(r"\s*(?:–|-)\s*")
TIME_SEGMENT_SPLIT = re.compile(r"\s*(?:&|,)\s*")
CLOCK_PATTERN = re.compile(r"(\d{1,2})(?::(\d{2}))?\s*(AM|PM)?", re.IGNORECASE)


@dataclass
class ItemRecord:
    id: str
    name: str
    category: str
    base_value: int
    rarity: str
    description: str
    season_tags: List[str]
    time_windows: List[Dict[str, float]]
    requirements: Optional[List[str]]


def slugify(name: str, prefix: str) -> str:
    safe = re.sub(r"[^0-9a-zA-Z]+", "_", name.lower()).strip("_")
    if not safe:
        safe = "item"
    if not safe.startswith(prefix):
        return f"{prefix}_{safe}"
    return safe


def to_number(cell_value) -> int:
    if cell_value is None:
        return 0
    if isinstance(cell_value, (int, float)):
        return int(cell_value)
    cleaned = str(cell_value).strip().replace(",", "")
    if not cleaned:
        return 0
    try:
        return int(float(cleaned))
    except ValueError:
        return 0


def determine_rarity(sell_price: int) -> str:
    last_label = "Common"
    for threshold, label in RARITY_THRESHOLDS:
        if sell_price < threshold:
            return last_label
        last_label = label
    return last_label


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def parse_time_segments(value: str) -> List[Dict[str, float]]:
    text = normalize_text(value)
    if not text or text.upper() == "NA":
        return []
    text = re.sub(r"\s*\(.*?\)", "", text).strip()
    if not text:
        return []
    if text.lower() == "all day":
        return [{"startHour": 0, "endHour": 24}]

    segments = [seg for seg in TIME_SEGMENT_SPLIT.split(text) if seg]
    windows: List[Dict[str, float]] = []
    for seg in segments:
        parts = TIME_RANGE_SPLIT.split(seg)
        if len(parts) != 2:
            continue
        start = parse_clock(parts[0])
        end = parse_clock(parts[1])
        if start is None or end is None:
            continue
        windows.append({"startHour": start, "endHour": end})
    return windows


def parse_clock(raw: str) -> Optional[float]:
    piece = normalize_text(raw)
    if not piece:
        return None
    match = CLOCK_PATTERN.match(piece)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2) or 0)
    ampm = match.group(3)
    if ampm:
        ampm = ampm.upper()
        if ampm == "AM":
            if hour == 12:
                hour = 0
        elif ampm == "PM":
            if hour != 12:
                hour += 12
    return hour + (minute / 60)


def derive_season_tags(month_values: Iterable[Tuple[int, str]]) -> List[str]:
    seasons = OrderedDict()
    for month_num, value in month_values:
        text = normalize_text(value)
        if not text or text.upper() == "NA":
            continue
        season = MONTH_TO_SEASON.get(month_num)
        if season:
            seasons[season] = True
    return list(seasons.keys())


def derive_time_windows(month_values: Iterable[str]) -> List[Dict[str, float]]:
    seen_strings: Dict[str, bool] = OrderedDict()
    for value in month_values:
        text = normalize_text(value)
        if not text or text.upper() == "NA":
            continue
        seen_strings[text] = True

    windows: List[Dict[str, float]] = []
    for text in seen_strings.keys():
        for window in parse_time_segments(text):
            if window not in windows:
                windows.append(window)
    return windows


def derive_requirements(where_how: str) -> Optional[List[str]]:
    text = normalize_text(where_how).lower()
    if not text:
        return None
    tags = OrderedDict()
    for keyword, tag in LOCATION_KEYWORDS:
        if keyword.lower() in text:
            tags[tag] = True
    return list(tags.keys()) or None


def format_luau(value, indent: int = 0) -> str:
    indent_str = "\t" * indent
    next_indent = "\t" * (indent + 1)

    if isinstance(value, dict):
        if not value:
            return "{}"
        lines = ["{"]
        for key, val in value.items():
            lines.append(f"{next_indent}{key} = {format_luau(val, indent + 1)},")
        lines.append(f"{indent_str}}}")
        return "\n".join(lines)
    if isinstance(value, list):
        if not value:
            return "{}"
        if all(not isinstance(item, (dict, list)) for item in value):
            return "{ " + ", ".join(format_luau(item, indent + 1) for item in value) + " }"
        lines = ["{"]
        for item in value:
            lines.append(f"{next_indent}{format_luau(item, indent + 1)},")
        lines.append(f"{indent_str}}}")
        return "\n".join(lines)
    if isinstance(value, str):
        escaped = value.replace("\\", "\\\\").replace('"', '\\"')
        return f'"{escaped}"'
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not value.is_integer():
            return f"{value:.2f}".rstrip("0").rstrip(".")
        return str(int(value)) if float(value).is_integer() else str(value)
    if value is None:
        return "nil"
    raise TypeError(f"Unsupported value: {value!r}")


def to_luau_module(items: Dict[str, ItemRecord]) -> str:
    lines = ["return {"]
    for item_id in sorted(items.keys()):
        record = items[item_id]
        lines.append(f"\t{item_id} = {{")
        fields: Sequence[Tuple[str, object]] = (
            ("id", record.id),
            ("name", record.name),
            ("category", record.category),
            ("baseValue", record.base_value),
            ("rarity", record.rarity),
            ("description", record.description),
            ("seasonTags", record.season_tags),
            ("timeWindows", record.time_windows),
            ("requirements", record.requirements),
        )
        for key, value in fields:
            if value is None:
                continue
            if isinstance(value, (list, dict)) and not value:
                continue
            lines.append(f"\t\t{key} = {format_luau(value, 2)},")
        lines.append("\t},")
    lines.append("}")
    return "\n".join(lines) + "\n"


def collect_items() -> Tuple[Dict[str, ItemRecord], List[str]]:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Cannot find {WORKBOOK_PATH}")

    workbook = load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)
    items: Dict[str, ItemRecord] = {}
    warnings: List[str] = []

    for sheet_name, category in CATEGORY_SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            warnings.append(f"Sheet '{sheet_name}' missing; skipping.")
            continue
        sheet = workbook[sheet_name]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        header = [str(col).strip() if col else "" for col in header_row]
        col_index = {name: idx for idx, name in enumerate(header)}

        missing_cols = [name for name, _ in MONTH_COLUMNS if name not in col_index]
        for col_name in ("Name", "Sell", "Description", "Where/How"):
            if col_name not in col_index:
                warnings.append(f"[{sheet_name}] column '{col_name}' missing.")
        if missing_cols:
            warnings.append(f"[{sheet_name}] missing month columns: {', '.join(missing_cols)}")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            name = normalize_text(row[col_index.get("Name", -1)]) if "Name" in col_index else ""
            if not name:
                continue
            prefix = category.lower()
            item_id = slugify(name, prefix)
            sell_price = to_number(row[col_index.get("Sell", -1)]) if "Sell" in col_index else 0
            description = normalize_text(row[col_index.get("Description", -1)]) if "Description" in col_index else ""
            where_how = normalize_text(row[col_index.get("Where/How", -1)]) if "Where/How" in col_index else ""

            month_values = []
            raw_times = []
            for month_name, month_num in MONTH_COLUMNS:
                value = row[col_index.get(month_name, -1)] if month_name in col_index else None
                month_values.append((month_num, value))
                raw_times.append(value)

            season_tags = derive_season_tags(month_values)
            time_windows = derive_time_windows(raw_times)
            requirements = derive_requirements(where_how)

            rarity = determine_rarity(sell_price)
            record = ItemRecord(
                id=item_id,
                name=name,
                category=category,
                base_value=sell_price,
                rarity=rarity,
                description=description,
                season_tags=season_tags,
                time_windows=time_windows,
                requirements=requirements,
            )
            if item_id in items:
                warnings.append(f"Duplicate item id '{item_id}' from sheet '{sheet_name}'.")
            items[item_id] = record

            if not time_windows:
                warnings.append(f"[{sheet_name}] '{name}' has no time windows parsed.")

    return items, warnings


def write_outputs(items: Dict[str, ItemRecord]) -> None:
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_LUA.parent.mkdir(parents=True, exist_ok=True)

    json_payload = {
        item_id: {
            "id": record.id,
            "name": record.name,
            "category": record.category,
            "baseValue": record.base_value,
            "rarity": record.rarity,
            "description": record.description,
            "seasonTags": record.season_tags,
            "timeWindows": record.time_windows,
            "requirements": record.requirements,
        }
        for item_id, record in items.items()
    }

    OUTPUT_JSON.write_text(json.dumps(json_payload, indent=2, ensure_ascii=False))
    OUTPUT_LUA.write_text(to_luau_module(items))


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Import ACNH spreadsheet slices.")
    parser.add_argument(
        "--json-only",
        action="store_true",
        help="Only emit JSON payload (skip Luau module).",
    )
    args = parser.parse_args(argv)

    items, warnings = collect_items()
    if not items:
        print("No items collected; aborting.")
        return 1

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(
        json.dumps(
            {
                item_id: {
                    "id": record.id,
                    "name": record.name,
                    "category": record.category,
                    "baseValue": record.base_value,
                    "rarity": record.rarity,
                    "description": record.description,
                    "seasonTags": record.season_tags,
                    "timeWindows": record.time_windows,
                    "requirements": record.requirements,
                }
                for item_id, record in items.items()
            },
            indent=2,
            ensure_ascii=False,
        )
    )

    if not args.json_only:
        OUTPUT_LUA.write_text(to_luau_module(items))

    print(f"Wrote {len(items)} items to {OUTPUT_JSON}")
    if not args.json_only:
        print(f"Wrote Luau module to {OUTPUT_LUA}")
    if warnings:
        print("Warnings:")
        for note in warnings[:25]:
            print(f" - {note}")
        if len(warnings) > 25:
            print(f"   ... {len(warnings) - 25} additional warnings.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
